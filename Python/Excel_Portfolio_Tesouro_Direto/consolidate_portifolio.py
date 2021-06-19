import pandas as pd
import openpyxl as excel
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.styles import NamedStyle
import datetime
import os

#  the directory containing the files to group the information should be like
#   path
#     |- 2020
#     |    |-2020-01.xlsx
#     |    |-...
#     |    |-2020-12.xlsx
#     |- 2021
#          |-2021-01.xlsx

# for debug, these options allow a better visualization of the dataframe in the debbug terminal.
pd.set_option('display.max_columns', None)
pd.set_option('display.width', None)
pd.set_option('display.max_colwidth', None)

def parse_excel(sheet, date, more_info, columns, info, header_depara):
    """Parse the excel data, get more_info data and return the dataframe contaning the columns 'Ticket', 'Date' and 'Value'.

    Args:
        sheet sheet (openpyxl): Object contaning the sheeet data.
        date (datetime): Data's reference date.
        more_info (dict): A dict to add information about
        columns (list): The columns name to be the header in the dataframe.
        info (list): The adicional info to get from the excel and add to more_info dict.
        header_depara (dict): A dict containing the excel header as key and the code header or variable name.

    Returns:
        DataFrame: DataFrame containing the data such as data, ticker, value.
    """
    data = []
    new_header = list(header_depara.keys())
    for row in sheet:                                                                                                                                   # get each row in the excel work sheet
        if row[new_header.index('maturity')].data_type != 'd':                                                                                      # cell data type 'd' is a datetime data, get the row if the first cell is an datetime value
            continue

        data.append([row[new_header.index(col)].value if col != 'Date' else date for col in columns])                                                 # add the columns in the columns list, it the column is Date, add the date that is not in the excel                                                                                                # append the values to the data matrix
        more_info.setdefault(row[new_header.index('Ticker')].value,
                             [row[new_header.index(col)].value if col != 'maturity' else row[new_header.index(col)].value.date() for col in info])

    return pd.DataFrame(data, columns=columns)


def get_data(path):
    """Access path directory, parse the excel file and return a DataFrame containing the excel data.

    Args:
        path (str): The full path directory to access the data file.
    """
    header_depara = {'Ticker': 'Titulo',
                     'maturity': 'Vencimento',
                     'purchased': 'Investido',
                     'Value': 'Bruto atual',
                     'Net Value': 'Liquído atual',
                     'amount': 'Total',
                     'blocked': 'Bloqueada'}

    columns=['Date', 'Ticker', 'Value']
    info = ['amount', 'purchased', 'maturity']
    more_info = {}                                                                                                                                      # ticker: [v_aplicado, qtd_aplicada, vencimento]

    years = os.listdir(path)                                                                                                                            # in the directory, for each year should have a folder, and in each folder should have an excel file for the reference month

    data = pd.DataFrame()
    for year in years:
        months = os.listdir(f'{path}\\{year}')                                                                                                          # access the folder for each year and get the list of files
        for month in months:
            sheet = excel.load_workbook(f'{path}\\{year}\\{month}', read_only=True).active                                                              # for each month (file in the folder) open the excel file and access the active sheet
            data = data.append(parse_excel(sheet, datetime.datetime.strptime(month, '%Y-%m.xlsx').date(), more_info, columns, info, header_depara))     # append the data form the sheet to the DataFrame containing all data

    data = data.reset_index().drop('index', axis=1)                                                                                                     # rest the dataframe index

    return data, more_info


def calculate_profitability(data, amount, purchased, maturity):
    """Calculate the profitability

    Args:
        data (DataFrame): DataFrame containing the historic data for a ticker.
        amount (float): The amout of the ticker purchased.
        purchased (float): The ticker's stock prices in the purchase data.
        maturity (datetime): The ticker maturity data.

    Returns:
        DataFrame: DataFrame containing the calculated values.
    """
    def profit_month(value, last_value):
        """Calculate the profit by month.

        Args:
            value (float): The ticker value for the month.
            last_value (float)): The ticker last month value or the purchased value.

        Returns:
            float: The calculated profit for the month (month value - last month value) / last month value
        """
        return (value - last_value ) / last_value

    def profit_all(value, purchased):
        """Calculate the profit by month.

        Args:
            value (float): The ticker value for the month.
            purchased (float)): The ticker purchased value.

        Returns:
            float: The calculated profit for the month (month value - purchased value) / purchased value
        """
        return (value - purchased ) / purchased

    data = data.copy().reset_index()                                                                                                                                  # copy the dataframe to avoid the SettingWithCopyWarning
    data.reset_index()                                                                                                                                  # reset the index
    for i, _ in data.iterrows():
        last = data.loc[i - 1, 'Value'] if i else purchased                                                                                            # get the last month value or the purchased value to calculate the profit

        data.loc[i, 'Profit month'] = profit_month(data.loc[i, 'Value'], last)
        data.loc[i, 'Profit all'] = profit_all(data.loc[i, 'Value'], purchased)

    return data


def plot_chart(sheet, header, chart_config, max_row):
    """Plot a chart for the data in the sheet.

    Args:
        sheet (openpyxl): The excel sheet for the ticker.
        header (list): The columns names in the sheet.
        chart_config (dict): A dict containing the chart title, and the axises label.
        max_row (inte): The data last row.
    """
    chart = ScatterChart()                                                                                                                              # defines the kind of chart (should be Scatter to define the axis x and y values)
    chart.title = chart_config['title']
    chart.x_axis.title = chart_config['x_axis']
    chart.y_axis.title = chart_config['y_axis']
    chart.height = chart_config['height']
    chart.width = chart_config['width']

    min_col = header.index('Value') + 1                                                                                                                 # the column index in the excel begin in 1, so to ajust the diference between the openpyxl and the list index add 1

    xvalues = Reference(sheet, min_col=header.index('Date') + 1, min_row=2, max_row=max_row)                                                                # min_row should be 2 to exclude the header form the reference
    for i in range(min_col + 1, len(header) + 1):
        values = Reference(sheet, min_col=i, min_row=2, max_row=max_row)                                                                                # create a reference to the column data
        series = Series(values, xvalues, title=sheet[1][i - 1].value)                                                                                   # assing the column data (y axsis) to the x values

        chart.series.append(series)                                                                                                                     # append the serie to the chart (shold be a multiple serie chart)

    sheet.add_chart(chart, "G1")


def format_row(sheet, row, number_format, header):
    """Format the row basead in the header.

    Args:
        sheet (openpyxl): The excel sheet for the ticker.
        row (int): The current row index.
        number_format (dict): A dict containing the number format rule, the key should be the column name (header) and value the pattern, like in excel.
        header (list): The columns name.
    """
    for col, name in enumerate(header):
        if number_format.get(name):                                                                                                                     # just if there is a rule for the column
            sheet[row][col].number_format  = number_format[name]                                                                                        # set the number format to the cell


def write_sheet(sheet, data, chart_config, number_format):
    """Write the data in to the sheet.

    Args:
        sheet (openpyxl): The excel sheet for the ticker.
        data (DataFrame): The data to insert in the excel sheet.
        chart_config (dict): The chart configuration, the key should be the property name and da value should be the velue to assign.
        number_format (dict): A dict containing the number format rule, the key should be the column name (header) and value the pattern, like in excel.
    """
    header =list(data.columns)                                                                                                                          # get the dataframe columns name into a list
    sheet.append(header)                                                                                                                                # add the header into the sheet

    for row_index, row in enumerate(data.values.tolist()):                                                                                                                    # convert the data frame to a matix
        sheet.append(row)                                                                                                                               # for each row add the data to the sheet
        format_row(sheet, row_index + 2, number_format, header)

    max_row = len(data) + 1                                                                                                                             # the data last row
    plot_chart(sheet, header, chart_config, max_row)


def data_to_excel(data, more_info, path):
    """Write the excel file containing a sheet for each ticker.

    Args:
        data (DataFrame): DataFrame containing the portfolio historic.
        more_info (dict): A dict containing the ticker as key and a list containing the application value, application amount and the ticker maturity
                          date as value
        path(str): The path to create the file.
    """
    wb = excel.Workbook()                                                                                                                               # inicialize the workbook
    chart_config = {'title': 'Profit',
                    'x_axis': 'Period - monthly',
                    'y_axis': 'Percent (%)',
                    'height': 10,
                    'width': 30}

    number_format = {'Date': 'yyyy-mm-dd',
                     'Value': '#,##0.00',
                     'Profit month': '0.00%',
                     'Profit all': '0.00%'}

    for ticker in sorted(list(set(data['Ticker']))):                                                                                                    # for each distinct Ticker in the dataframe
        sheet = wb.create_sheet()                                                                                                                       # create a sheet for each ticker
        sheet.title = ticker.replace('Juros Semestrais', 'JS')                                                                                          # rename the sheet as the ticker name, replace 'Juros Semestrais', 'JS' because the sheet name in the excel has a limit in the number of the characters
        calculated = calculate_profitability(data[data['Ticker'] == ticker], *more_info[ticker])                                                        # get the rows in the dataframe wher the column Ticker is equal to the ticker in the loop

        write_sheet(sheet, calculated[['Date', 'Ticker', 'Value', 'Profit month', 'Profit all']], chart_config, number_format)                          # get the DataFrame columns in the order

    wb.remove(wb['Sheet'])

    wb.save(path)


path_data = r'C:\Users\carol\Downloads\Carteira\Extratos'
path_excel = r'C:\Users\carol\Downloads\Carteira\Consolidated.xlsx'
data, more_info = get_data(path_data)

data_to_excel(data, more_info, path_excel)
